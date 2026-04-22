import asyncio
import serial
import serial.tools.list_ports
import websockets
import json
import threading
import queue
import argparse
import random
import time
import math
import collections
import http.server
import socket
import socketserver
import signal
import atexit
import os
import subprocess
import sys
from typing import List, Dict
import logging
import csv
from datetime import datetime, timezone
from pathlib import Path

# Configuration
BAUD_RATE = 460800
WS_PORT = 3941  # WebSocket port for UIs to connect
HTTP_PORT = 8080  # Static file server port
UDP_GROUP = "239.255.0.1"
UDP_PORT = 5005
FORCE_SERIAL_PORT = "COM5"  # Hardcoded per request

class TestDataGenerator:
    def __init__(self, test_data_file="test_data.json", scenario="normal_operation"):
        self.test_data = {}
        self.current_scenario = scenario
        self.scenario_index = 0
        self.available_scenarios = []
        self.noise_levels = {}
        self.update_rates = {}
        self.last_values = {}
        self.start_time = time.time()
        
        try:
            with open(test_data_file, 'r') as f:
                data = json.load(f)
                self.test_data = data.get("test_scenarios", {})
                self.noise_levels = data.get("noise_levels", {})
                self.update_rates = data.get("update_rates", {})
                self.available_scenarios = list(self.test_data.keys())
                
                if scenario not in self.available_scenarios and self.available_scenarios:
                    self.current_scenario = self.available_scenarios[0]
                    
                print(f"Test data loaded. Available scenarios: {self.available_scenarios}")
                print(f"Using scenario: {self.current_scenario}")
                
        except FileNotFoundError:
            print("Warning: test_data.json not found, using default values")
            self._create_default_data()
        except Exception as e:
            print(f"Error loading test data: {e}")
            self._create_default_data()
    
    def _create_default_data(self):
        """Create basic test data if file is missing"""
        self.test_data = {
            "normal_operation": {
                "pt_channels": {str(i): {"value": 100 + i * 10, "trend": "stable"} for i in range(1, 17)},
                "lc_channels": {str(i): {"value": 200 + i * 50, "trend": "stable"} for i in range(1, 7)},
                "tc_channels": {str(i): {"value": 300 + i * 100, "trend": "stable"} for i in range(1, 7)}
            }
        }
        self.available_scenarios = ["normal_operation"]
        self.current_scenario = "normal_operation"
        self.noise_levels = {"pt_noise": 2.5, "lc_noise": 5.0, "tc_noise": 10.0}
        self.update_rates = {"pt_fast": 0.5, "lc_fast": 0.5, "tc_fast": 1.0}
    
    def cycle_scenario(self):
        """Cycle to the next test scenario"""
        if len(self.available_scenarios) > 1:
            self.scenario_index = (self.scenario_index + 1) % len(self.available_scenarios)
            self.current_scenario = self.available_scenarios[self.scenario_index]
            print(f"Switched to scenario: {self.current_scenario}")
            return True
        return False
    
    def generate_sensor_value(self, sensor_type, channel_id, base_value, trend="stable"):
        """Generate a realistic sensor value with noise and trends"""
        noise_key = f"{sensor_type}_noise"
        noise_level = self.noise_levels.get(noise_key, 5.0)
        
        # Add noise
        noise = random.uniform(-noise_level, noise_level)
        
        # Add time-based variation
        elapsed = time.time() - self.start_time
        
        # Trend effects
        trend_factor = 0
        if trend == "increasing":
            trend_factor = math.sin(elapsed * 0.1) * noise_level * 0.5
        elif trend == "decreasing":
            trend_factor = -math.sin(elapsed * 0.1) * noise_level * 0.5
        elif trend == "oscillating":
            trend_factor = math.sin(elapsed * 0.2) * noise_level
        
        # Smooth value changes
        key = f"{sensor_type}_{channel_id}"
        if key in self.last_values:
            # Smooth transition
            target_value = base_value + noise + trend_factor
            current_value = self.last_values[key]
            smooth_factor = 0.1  # How quickly values change
            new_value = current_value + (target_value - current_value) * smooth_factor
        else:
            new_value = base_value + noise + trend_factor
        
        self.last_values[key] = new_value
        return new_value  # Allow negative values for realistic sensor data
    
    def get_csv_packet(self, sensor_type):
        """Generate a CSV packet for all channels of a given sensor type"""
        scenario_data = self.test_data.get(self.current_scenario, {})
        channels_data = scenario_data.get(f"{sensor_type}_channels", {})
        
        values = []
        if sensor_type == "pt":
            # PT channels 0-15
            for i in range(16):
                channel_data = channels_data.get(str(i), {"value": 0.0, "trend": "stable"})
                value = self.generate_sensor_value(sensor_type, i, channel_data["value"], channel_data["trend"])
                values.append(f"p{value:.6f}")
        elif sensor_type == "lc" or sensor_type == "tc":
            # LC channels 0-5 + TC channels 0-5 in one 't' packet
            # First 6 are LC (index 0-5), next 6 are TC (index 0-5)
            for i in range(6):
                # LC channels
                lc_data = scenario_data.get("lc_channels", {}).get(str(i), {"value": 0.0, "trend": "stable"})
                lc_value = self.generate_sensor_value("lc", i, lc_data["value"], lc_data["trend"])
                values.append(f"t{lc_value:.6f}")
            
            for i in range(6):
                # TC channels
                tc_data = scenario_data.get("tc_channels", {}).get(str(i), {"value": 0.0, "trend": "stable"})
                tc_value = self.generate_sensor_value("tc", i, tc_data["value"], tc_data["trend"])
                values.append(f"t{tc_value:.6f}")
        elif sensor_type == "dc":
            # DC/Solenoid channels 0-11
            for i in range(12):
                channel_data = channels_data.get(str(i), {"value": 0.0, "trend": "stable"})
                value = self.generate_sensor_value(sensor_type, i, channel_data["value"], channel_data["trend"])
                values.append(f"s{value:.6f}")
        
        return ",".join(values)

class SerialProcessor:
    def __init__(self, test_mode=False, test_scenario="normal_operation", udp_enable=True, udp_group=UDP_GROUP, udp_port=UDP_PORT):
        self.ser = None
        self.connected = False
        self.test_mode = test_mode
        self.data_queue = queue.Queue()  # For incoming serial data
        self.clients: List[websockets.WebSocketServerProtocol] = []  # Connected WebSocket clients
        self.loop = asyncio.get_event_loop()
        self.udp_enable = udp_enable
        self.udp_group = udp_group
        self.udp_port = udp_port
        self.udp_sock = None
        if self.udp_enable:
            try:
                self.udp_sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM, socket.IPPROTO_UDP)
                # Set a small TTL for local network
                self.udp_sock.setsockopt(socket.IPPROTO_IP, socket.IP_MULTICAST_TTL, 1)
                # Enable loopback so the host can also receive if needed
                self.udp_sock.setsockopt(socket.IPPROTO_IP, socket.IP_MULTICAST_LOOP, 1)
                print(f"UDP multicast enabled: group={self.udp_group} port={self.udp_port}")
            except Exception as e:
                print(f"Failed to initialize UDP multicast: {e}")
                self.udp_enable = False
        
        # Processing / calibration configuration (applies to outgoing data)
        self.apply_pt_calibration: bool = True
        self.pt_min_ma: float = 4.0
        self.pt_max_ma: float = 20.0
        self.pt_shunt_ohms: float = 47.0  # default shunt used on hardware
        # Optional per-channel effective shunt ohms (computed at tare)
        self.pt_shunt_ohms_list: List[float] = [self.pt_shunt_ohms for _ in range(16)]
        # PandaV2 scuffed-bangbang firmware emits raw voltage across the PT
        # shunt (~0.7 V at idle), not mA. Stay in 'volts' mode so the server's
        # auto-shunt calibration converts V -> mA against the 4 mA baseline.
        self.pt_input_mode: str = "volts"  # auto | ma | volts
        self.pt_meta: List[Dict] = self._load_pt_metadata()
        # Per-channel tare offsets in engineering units (psi)
        self.pt_offsets: List[float] = [0.0 for _ in range(16)]
        # Last seen raw PT currents (mA) for tare capture
        self._last_pt_mA: List[float] = [0.0 for _ in range(16)]
        self._last_pt_raw: List[float] = [0.0 for _ in range(16)]
        # If V1 emits uppercase 'P' PT PSI rows, prefer those for outgoing PT values.
        self._prefer_v1_pt_psi: bool = False
        self._last_v1_pt_psi_ts: float = 0.0
        # Debug: echo raw serial bus lines (off by default — very noisy)
        self.debug_print_serial: bool = False
        # Throttle for console RAW printing (seconds between prints)
        self.debug_raw_throttle_sec: float = 1.0
        self._last_raw_print: float = 0.0
        # Throttle noisy PT preview prints (seconds between prints)
        self._last_pt_preview_print: float = 0.0
        self._pt_preview_throttle_sec: float = 1.0
        # Cap websocket send cadence to avoid backpressure/latency stalls.
        self.ws_max_hz: float = 40.0
        self._last_ws_send_ts: float = 0.0
        # Auto-estimate shunt from ambient voltage so ambient ≈ 4 mA in volts mode
        self.auto_shunt_v_ambient: bool = True
        self._ambient_v_sum: float = 0.0
        self._ambient_v_count: int = 0
        # DC on-current threshold (Amps). Default 0.100 A per request.
        self.dc_threshold: float = 0.100

        # Test mode setup
        if self.test_mode:
            self.test_generator = TestDataGenerator(scenario=test_scenario)
            self.test_running = False
            print(f"Running in TEST MODE with scenario: {test_scenario}")
        else:
            self.test_generator = None
        
        # CSV logging
        self.logging_enabled = False
        self.log_file = None
        self.log_writer = None
        self.log_filename = None
        self.log_lock = threading.Lock()
        self.log_data_count = 0
        self.log_start_time = None
        self.current_data = {'pt': {}, 'lc': {}, 'tc': {}, 'dc': {}, 'dc_state': {}}
        # Command logging (DC commands and any others sent via WS)
        self.cmd_log_file = None
        self.cmd_log_writer = None
        self.cmd_log_filename = None
        # Bang-bang config cache + latest firmware runtime heartbeat mirror.
        self.bb_configs = {
            "lox":  {"setpoint": 200.0, "deadband": 10.0, "wait_ms": 500, "max_open_ms": 0,
                     "vent_trigger": 0.0, "vent_auto": False,
                     "mdot_target": 0.0, "sp_min": 0.0, "sp_max": 0.0, "gain": 0.0, "rho": 0.0, "enable": False},
            "fuel": {"setpoint": 200.0, "deadband": 10.0, "wait_ms": 500, "max_open_ms": 0,
                     "vent_trigger": 0.0, "vent_auto": False,
                     "mdot_target": 0.0, "sp_min": 0.0, "sp_max": 0.0, "gain": 0.0, "rho": 0.0, "enable": False},
        }
        self.bb_runtime = {
            "l": {"state": "OFF", "press": False, "vent": False, "pressure": 0.0},
            "f": {"state": "OFF", "press": False, "vent": False, "pressure": 0.0},
        }
        self.device_messages = collections.deque(maxlen=300)
        # Preset sequences (e.g., GOX purge, hotfire) persisted to disk
        base_dir = os.path.dirname(os.path.abspath(__file__))
        self.preset_file = Path(base_dir) / "configs" / "preset_sequences.json"
        self.presets = {"gox": None, "hotfire": None}
        self._load_presets()

    def _load_pt_metadata(self) -> List[Dict]:
        """Load PT metadata (names/units/range) from sensor_config.xlsx for panda system.
        Returns a 16-element list mapped by channel index.
        """
        try:
            import openpyxl
            
            base_dir = os.path.dirname(os.path.abspath(__file__))
            xlsx_path = os.path.join(base_dir, '..', 'sensor_config.xlsx')
            if not os.path.exists(xlsx_path):
                xlsx_path = os.path.join(base_dir, 'sensor_config.xlsx')
            
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb['PT_Sensors']
            
            headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            col = {h: i for i, h in enumerate(headers) if h}
            
            mapped: List[Dict] = [{} for _ in range(16)]
            for row in ws.iter_rows(min_row=2, values_only=True):
                system = row[col.get('system', 0)]
                if not system or str(system).strip().lower() != 'panda':
                    continue
                
                try:
                    channel = int(row[col['channel']])
                except (ValueError, KeyError, TypeError):
                    continue
                
                if 0 <= channel < 16:
                    mapped[channel] = {
                        "id": channel + 1,
                        "name": str(row[col.get('name', -1)] or f'PT{channel + 1}'),
                        "short_name": str(row[col.get('short_name', -1)] or f'pt{channel + 1}'),
                        "units": str(row[col.get('units', -1)] or 'psi'),
                        "range": [0, 1500],
                    }
            
            wb.close()
            
            for i in range(16):
                if not mapped[i]:
                    mapped[i] = {"id": i + 1, "range": [0, 1500], "units": "psi"}

            # Overlay authoritative ranges from pt_channels.json when available.
            try:
                pt_cfg_path = os.path.join(base_dir, 'configs', 'pt_channels.json')
                if os.path.exists(pt_cfg_path):
                    with open(pt_cfg_path, 'r', encoding='utf-8') as f:
                        pt_cfg = json.load(f)
                    for ch in pt_cfg.get('channels', []):
                        pt = ch.get('pt') or {}
                        idx = int(pt.get('raw_index', pt.get('id', 0))) - 1
                        rng = pt.get('range')
                        if 0 <= idx < 16 and isinstance(rng, list) and len(rng) == 2:
                            mapped[idx]['range'] = [float(rng[0]), float(rng[1])]
                            if pt.get('name'):
                                mapped[idx]['name'] = str(pt['name'])
                            if pt.get('short_name'):
                                mapped[idx]['short_name'] = str(pt['short_name'])
            except Exception as e:
                print(f"PT range overlay from pt_channels.json failed: {e}")

            try:
                r0 = mapped[0].get('range', [0, 1500])
                print(f"PT0 range loaded: {r0}")
            except Exception:
                pass
            return mapped
        except Exception as e:
            print(f"Failed to load PT metadata from xlsx: {e}")
            return [{"id": i + 1, "range": [0, 1500], "units": "psi"} for i in range(16)]

    def _calibrate_pt_values(self, values: List[float]) -> List[float]:
        """Convert PT readings to engineering units using 4-20mA mapping and channel ranges.
        - If pt_input_mode == 'ma', treat values as milliamps
        - If pt_input_mode == 'volts', treat as volts across shunt and convert to mA via shunt
        - If 'auto', infer based on magnitude: <=2V -> volts; <=30 -> mA; otherwise already in units
        """
        if not values:
            return values
        # Reuse unified conversion so polarity and mode handling are consistent
        vals_mA: List[float] = self._to_mA(values)
        if not vals_mA:
            # Firmware always emits mA; fall back to treating values as mA directly
            # rather than forwarding raw numbers as engineering units.
            vals_mA = list(values)

        calibrated: List[float] = []
        for idx, mA in enumerate(vals_mA):
            # Normalize to 0..1 from 4..20 mA
            norm = (mA - self.pt_min_ma) / (self.pt_max_ma - self.pt_min_ma)
            norm = max(0.0, min(1.0, norm))
            meta = self.pt_meta[idx] if idx < len(self.pt_meta) else {}
            rng = meta.get('range', [0, 1000])
            try:
                min_eng, max_eng = float(rng[0]), float(rng[1])
            except Exception:
                min_eng, max_eng = 0.0, 1000.0
            eng_val = min_eng + norm * (max_eng - min_eng)
            # Apply tare offset in engineering units (psi)
            try:
                eng_val -= float(self.pt_offsets[idx])
            except Exception:
                pass
            calibrated.append(eng_val)
        return calibrated

    def _to_mA(self, values: List[float]) -> List[float]:
        """Convert incoming PT values to milliamps according to pt_input_mode (or inference)."""
        if not values:
            return []
        mode = self.pt_input_mode
        try:
            vmin, vmax = min(values), max(values)
            vavg = sum(values) / len(values)
        except Exception:
            vmin, vmax, vavg = 0.0, 0.0, 0.0
        if mode == "volts":
            out: List[float] = []
            for idx, v in enumerate(values):
                pol_i = -1.0 if v < 0.0 else 1.0
                r_i = None
                try:
                    if isinstance(self.pt_shunt_ohms_list, list) and idx < len(self.pt_shunt_ohms_list):
                        r_i = float(self.pt_shunt_ohms_list[idx])
                except Exception:
                    r_i = None
                if not r_i or r_i <= 0.0:
                    r_i = self.pt_shunt_ohms
                out.append(((pol_i * v) / max(r_i, 0.001)) * 1000.0)
            return out
        if mode == "ma":
            return values[:]
        # auto
        abs_max = max(abs(vmin), abs(vmax))
        if abs_max <= 2.0:
            out: List[float] = []
            for idx, v in enumerate(values):
                pol_i = -1.0 if v < 0.0 else 1.0
                r_i = None
                try:
                    if isinstance(self.pt_shunt_ohms_list, list) and idx < len(self.pt_shunt_ohms_list):
                        r_i = float(self.pt_shunt_ohms_list[idx])
                except Exception:
                    r_i = None
                if not r_i or r_i <= 0.0:
                    r_i = self.pt_shunt_ohms
                out.append(((pol_i * v) / max(r_i, 0.001)) * 1000.0)
            return out
        # Accept slight negatives (idle / unconnected channel noise) as mA too.
        if -2.0 <= vmin <= 30.0 and -2.0 <= vmax <= 30.0:
            return values[:]
        return []

    def _compute_eng_from_ma_no_tare(self, vals_mA: List[float]) -> List[float]:
        out: List[float] = []
        for idx, mA in enumerate(vals_mA):
            norm = (mA - self.pt_min_ma) / (self.pt_max_ma - self.pt_min_ma)
            norm = max(0.0, min(1.0, norm))
            meta = self.pt_meta[idx] if idx < len(self.pt_meta) else {}
            rng = meta.get('range', [0, 1000])
            try:
                min_eng, max_eng = float(rng[0]), float(rng[1])
            except Exception:
                min_eng, max_eng = 0.0, 1000.0
            out.append(min_eng + norm * (max_eng - min_eng))
        return out


    def _process_outgoing_data_and_meta(self, raw: str):
        """Apply processing/calibration to outgoing CSV packets and return (text, dc_meta, device_msgs).
        dc_meta is a dict with keys: timestamp, currents, states; or None if not applicable.
        """
        try:
            lines = [ln for ln in raw.split('\n') if ln.strip()]
            out_lines: List[str] = []
            dc_meta = None
            device_msgs: List[Dict] = []
            for line in lines:
                if ',' not in line:
                    # Firmware status / event lines (non-CSV)
                    if line.startswith("BB:"):
                        try:
                            # BB:<side>:<state>:<press01>:<vent01>:<pressure_psi>
                            parts = line.split(':')
                            if len(parts) >= 5:
                                bus = (parts[1] or '').lower()
                                if bus in self.bb_runtime:
                                    self.bb_runtime[bus]["state"] = parts[2]
                                    self.bb_runtime[bus]["press"] = parts[3] == '1'
                                    self.bb_runtime[bus]["vent"] = parts[4] == '1'
                                    if len(parts) > 5:
                                        self.bb_runtime[bus]["pressure"] = float(parts[5])
                        except Exception:
                            pass
                    if line.startswith(("EVT:", "BB_ERROR:", "CMD_ERROR:", "Arming!", "Disarming!", "SEQ_")):
                        device_msgs.append({
                            "type": "device_message",
                            "message": line,
                            "timestamp": datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
                        })
                    out_lines.append(line)
                    continue
                tokens = [tok.strip() for tok in line.split(',')]
                if not tokens:
                    continue
                id_char = tokens[0][0] if tokens[0] else ''
                # Parse numeric portion from each token
                def _to_float(tok: str) -> float:
                    try:
                        s = ''.join(ch for ch in tok if ch.isdigit() or ch in ['-', '+', '.'])
                        return float(s) if s not in ['', '+', '-'] else 0.0
                    except Exception:
                        return 0.0

                # Drop V1-scaled 'P' PSI rows entirely — the server re-computes PSI
                # from the lowercase 'p' mA stream so raw/mA/psi stay consistent.
                if id_char == 'P':
                    continue
                if id_char == 'p' and self.apply_pt_calibration:
                    # Board sends currents in mA for PTs; tokens look like p6.500, p6.520 ...
                    numeric = [_to_float(tok) for tok in tokens]
                    try:
                        if numeric:
                            # Store latest raw PT values for taring (pad to 16)
                            raw16 = (numeric + [0.0]*16)[:16]
                            self._last_pt_raw = raw16
                    except Exception:
                        pass
                    # Auto-estimate shunt from ambient (voltage mode): target 4 mA at idle
                    if self.pt_input_mode == "volts" and self.auto_shunt_v_ambient:
                        try:
                            if numeric:
                                avg_abs_v = sum(abs(v) for v in numeric[:min(16, len(numeric))]) / float(min(16, len(numeric)))
                                self._ambient_v_sum += max(0.0, avg_abs_v)
                                self._ambient_v_count += 1
                                if self._ambient_v_count >= 5:
                                    mean_v = self._ambient_v_sum / self._ambient_v_count
                                    # R = V / I, with I = 0.004 A
                                    est_r = max(1.0, mean_v / 0.004)
                                    self.pt_shunt_ohms = est_r
                                    self.auto_shunt_v_ambient = False
                                    print(f"Auto-set pt_shunt_ohms≈{est_r:.2f}Ω from ambient {mean_v:.4f} V -> 4 mA")
                        except Exception:
                            pass
                    # Convert to mA for tare capture based on input mode
                    try:
                        mA_vals = self._to_mA(numeric)
                        if mA_vals:
                            if len(mA_vals) < 16:
                                mA_vals = mA_vals + [0.0]*(16-len(mA_vals))
                            self._last_pt_mA = mA_vals[:16]
                            # Periodic debug print of first channels in mA
                            now_dbg = time.time()
                            if (now_dbg - self._last_pt_preview_print) >= self._pt_preview_throttle_sec:
                                try:
                                    preview = ', '.join(f"{v:.3f}mA" for v in mA_vals[:4])
                                    print(f"PT mA preview: {preview}")
                                    self._last_pt_preview_print = now_dbg
                                except Exception:
                                    pass
                    except Exception:
                        pass
                    calibrated = self._calibrate_pt_values(numeric)
                    rebuilt = ','.join([f"p{v:.6f}" for v in calibrated])
                    out_lines.append(rebuilt)
                    # Store for logging (PT channels 1-16)
                    for i, val in enumerate(calibrated[:16], 1):
                        self.current_data['pt'][i] = val
                    # Periodic sanity print of the mA -> psi mapping for ch0.
                    try:
                        now_dbg2 = time.time()
                        if int(now_dbg2) % 2 == 0 and self._last_raw_print != int(now_dbg2):
                            self._last_raw_print = int(now_dbg2)
                            rng0 = (self.pt_meta[0] or {}).get('range', [0, 1500]) if self.pt_meta else [0, 1500]
                            ma0 = self._last_pt_mA[0] if self._last_pt_mA else 0.0
                            psi0 = calibrated[0] if calibrated else 0.0
                            print(f"PT0 map: {ma0:.3f}mA -> {psi0:.2f}psi (range {rng0}, mode={self.pt_input_mode})")
                    except Exception:
                        pass
                elif id_char == 'l':
                    # Legacy: load cell data on 'l' line
                    numeric = [_to_float(tok) for tok in tokens]
                    out_lines.append(line)
                    for i, val in enumerate(numeric[:6], 1):
                        self.current_data['lc'][i] = val
                elif id_char == 't':
                    numeric = [_to_float(tok) for tok in tokens]
                    out_lines.append(line)
                    num_lc = min(8, len(numeric))
                    for i, val in enumerate(numeric[:num_lc], 1):
                        self.current_data['lc'][i] = val
                    for i, val in enumerate(numeric[num_lc:num_lc + 8], 1):
                        self.current_data['tc'][i] = val
                elif id_char == 's':
                    currents = [_to_float(tok) for tok in tokens]
                    states = [val >= getattr(self, 'dc_threshold', 0.1) for val in currents]
                    dc_meta = {
                        "timestamp": datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
                        "currents": currents,
                        "states": states,
                    }
                    out_lines.append(line)
                    # Store for logging (DC channels 1-16)
                    for i, val in enumerate(currents[:16], 1):
                        self.current_data['dc'][i] = val
                    for i, st in enumerate(states[:16], 1):
                        self.current_data['dc_state'][i] = 1 if st else 0
                else:
                    out_lines.append(line)
            return '\n'.join(out_lines), dc_meta, device_msgs
        except Exception:
            return raw, None, []

    async def connect(self, port: str) -> Dict:
        if self.connected:
            # If already connected to the requested port, acknowledge without reconnecting
            try:
                current = getattr(self.ser, 'port', None) if self.ser else None
            except Exception:
                current = None
            if current and str(current).lower() == FORCE_SERIAL_PORT.lower():
                logging.info(f"Already connected to {current}; ignoring duplicate connect request")
                return {"success": True, "message": f"Already connected to {current}"}
            # Otherwise, switch ports
            self.disconnect()
            
        if self.test_mode:
            # In test mode, simulate connection
            self.connected = True
            self.test_running = True
            threading.Thread(target=self.generate_test_data, daemon=True).start()
            return {"success": True, "message": f"Connected to TEST MODE ({FORCE_SERIAL_PORT})"}
        else:
            # Normal serial connection
            try:
                self.ser = serial.Serial(FORCE_SERIAL_PORT, BAUD_RATE, timeout=2)
                self.connected = True
                threading.Thread(target=self.read_serial, daemon=True).start()
                return {"success": True, "message": f"Connected to {FORCE_SERIAL_PORT}"}
            except Exception as e:
                return {"success": False, "message": str(e)}

    def disconnect(self):
        self.connected = False
        if self.test_mode:
            self.test_running = False
        if self.ser:
            self.ser.close()

    def read_serial(self):
        lines_read = 0
        bytes_read = 0
        overflow_count = 0
        last_report = time.time()
        _buf = bytearray()

        def _format_preview(buf: bytes, n: int = 96) -> str:
            """Return 'hex: .. .. .. | ascii: \"...\"' for the last n bytes."""
            tail = bytes(buf[-n:])
            hex_part = ' '.join(f"{b:02x}" for b in tail)
            ascii_part = ''.join(chr(b) if 0x20 <= b <= 0x7e else '.' for b in tail)
            return f"hex: {hex_part}  |  ascii: \"{ascii_part}\""

        def _diagnose(buf: bytes) -> str:
            """Try several alternative byte interpretations and report which one
            actually contains newlines. The winner tells us how the link is broken."""
            candidates = {
                "as-is":            buf,
                "XOR 0xFF (inv)":   bytes(b ^ 0xFF for b in buf),
                "high-bit-strip":   bytes(b & 0x7F for b in buf),
                "high-bit-set":     bytes(b | 0x80 for b in buf),
                "bit-reverse":      bytes(int(f"{b:08b}"[::-1], 2) for b in buf),
            }
            scored = []
            for name, data in candidates.items():
                nl = data.count(0x0A) + data.count(0x0D)
                printable = sum(1 for b in data if 0x20 <= b <= 0x7E)
                scored.append((name, nl, printable, len(data)))
            best = max(scored, key=lambda r: (r[1], r[2]))
            lines = [f"  - {n}: newlines={nl} printable={pr}/{tot}"
                     for (n, nl, pr, tot) in scored]
            verdict = ""
            if best[1] > 0 and best[0] != "as-is":
                verdict = (f"\n  >>> '{best[0]}' contains newlines — "
                           f"link likely needs that transformation upstream.")
            return "Alt-decode scan:\n" + "\n".join(lines) + verdict

        while self.connected:
            try:
                waiting = self.ser.in_waiting
                chunk = self.ser.read(max(1, waiting))
                if not chunk:
                    now = time.time()
                    if now - last_report >= 5.0:
                        logging.info("Serial: no data flowing (%.1f s silence)",
                                     now - last_report)
                        last_report = now
                    continue

                bytes_read += len(chunk)
                _buf.extend(chunk)

                while b'\n' in _buf:
                    idx = _buf.index(b'\n')
                    raw_line = bytes(_buf[:idx])
                    del _buf[:idx + 1]
                    line = raw_line.decode('ascii', errors='replace').strip()
                    if not line:
                        continue

                    if self.debug_print_serial:
                        try:
                            now = time.time()
                            if (now - self._last_raw_print) >= max(
                                0.0, self.debug_raw_throttle_sec
                            ):
                                print(f"RAW: {line}")
                                self._last_raw_print = now
                        except Exception:
                            pass

                    self.data_queue.put(line)
                    lines_read += 1

                if len(_buf) > 16384:
                    overflow_count += 1
                    sample = bytes(_buf[-128:])
                    raw_ascii  = sum(1 for b in sample if 0x20 <= b <= 0x7E)
                    flip_ascii = sum(1 for b in sample if 0x20 <= (b ^ 0xFF) <= 0x7E)
                    hint = ""
                    if flip_ascii > raw_ascii + 4:
                        hint = " Bytes look bit-inverted → check RS-485 A/B polarity or firmware TXINV."
                    elif raw_ascii < 8:
                        hint = " Almost no printable ASCII → wrong baud or non-ASCII framing."

                    logging.warning(
                        "Serial overflow: %dB without newline.%s\n  %s\n  %s",
                        len(_buf), hint,
                        _format_preview(_buf, 96),
                        _diagnose(bytes(_buf[:512])),
                    )
                    _buf.clear()

                now = time.time()
                if now - last_report >= 5.0:
                    elapsed = now - last_report
                    kbps = bytes_read / elapsed / 1000
                    if lines_read > 0:
                        logging.info("Serial: %.1f lines/s  %.1f kB/s  overflows=%d",
                                     lines_read / elapsed, kbps, overflow_count)
                    else:
                        logging.info("Serial: 0 lines parsed  %.1f kB/s raw  overflows=%d  "
                                     "(no newlines — RS-485 polarity / baud?)",
                                     kbps, overflow_count)
                    lines_read = 0
                    bytes_read = 0
                    overflow_count = 0
                    last_report = now

            except Exception:
                pass
    
    def generate_test_data(self):
        """Generate test telemetry data in CSV format"""
        cycle_count = 0
        while self.test_running and self.connected:
            try:
                # Generate PT packet (16 channels)
                pt_packet = self.test_generator.get_csv_packet("pt")
                self.data_queue.put(pt_packet)
                time.sleep(0.5)
                
                if not self.test_running:
                    break
                
                # Generate combined LC+TC packet (12 channels total)
                lc_tc_packet = self.test_generator.get_csv_packet("lc")  # This handles both LC and TC
                self.data_queue.put(lc_tc_packet)
                time.sleep(0.5)
                
                if not self.test_running:
                    break
                
                # Generate DC/Solenoid packet (12 channels)
                dc_packet = self.test_generator.get_csv_packet("dc")
                self.data_queue.put(dc_packet)
                time.sleep(0.5)
                
                # Cycle scenarios every 20 seconds
                cycle_count += 1
                if cycle_count % 40 == 0:  # 40 iterations ≈ 20 seconds (3 packets * 0.5s each + delays)
                    if self.test_generator.cycle_scenario():
                        print(f"Auto-switched to scenario: {self.test_generator.current_scenario}")
                
            except Exception as e:
                print(f"Error generating test data: {e}")
                time.sleep(1)

    async def send_command(self, command: str) -> Dict:
        if not self.connected:
            result = {"success": False, "message": "Not connected"}
            try:
                self._write_command_entry(command, success=False, result=result.get("message", ""))
            except Exception:
                pass
            return result
            
        if self.test_mode:
            # Handle special test commands
            if command == "test_cycle":
                if self.test_generator.cycle_scenario():
                    result = {"success": True, "message": f"Switched to scenario: {self.test_generator.current_scenario}"}
                else:
                    result = {"success": True, "message": "Only one scenario available"}
            elif command.startswith("test_scenario_"):
                scenario = command.replace("test_scenario_", "")
                if scenario in self.test_generator.available_scenarios:
                    self.test_generator.current_scenario = scenario
                    result = {"success": True, "message": f"Switched to scenario: {scenario}"}
                else:
                    result = {"success": False, "message": f"Unknown scenario: {scenario}"}
            else:
                # Simulate normal command acknowledgment
                result = {"success": True, "message": f"TEST MODE - Simulated: {command}"}
            try:
                self._write_command_entry(command, success=bool(result.get("success")), result=str(result.get("message", "")))
            except Exception:
                pass
            return result
        else:
            # Normal serial command
            try:
                self.ser.write((command + '\n').encode('utf-8'))
                self.ser.flush()
                result = {"success": True, "message": f"Sent: {command}"}
            except Exception as e:
                result = {"success": False, "message": str(e)}
            try:
                self._write_command_entry(command, success=bool(result.get("success")), result=str(result.get("message", "")))
            except Exception:
                pass
            return result

    async def broadcast_data(self):
        ingest_count = 0
        last_print = time.time()
        send_interval = 1.0 / max(1.0, float(getattr(self, "ws_max_hz", 40.0)))
        while True:
            processed_any = False
            latest_processed = None
            pending_device_msgs: List[Dict] = []
            try:
                while True:
                    try:
                        raw_data = self.data_queue.get_nowait()
                    except queue.Empty:
                        break

                    processed_any = True
                    data, dc_meta, device_msgs = self._process_outgoing_data_and_meta(raw_data)
                    ingest_count += 1
                    latest_processed = (data, raw_data, dc_meta)
                    if device_msgs:
                        # Cap burst event fan-out per loop.
                        pending_device_msgs.extend(device_msgs[:64])
                    now = time.time()
                    if now - last_print >= 2.0:
                        try:
                            preview = (raw_data[:120] + '...') if len(raw_data) > 120 else raw_data
                            rate = ingest_count / (now - last_print)
                            logging.info(f"WS ingest rate: {rate:.1f} lines/s, clients={len(self.clients)}; raw preview: {preview}")
                        except Exception:
                            pass
                        ingest_count = 0
                        last_print = now

                    now_send = time.time()
                    should_send = (
                        latest_processed is not None
                        and self.clients
                        and (now_send - self._last_ws_send_ts) >= send_interval
                    )
                    if should_send:
                        data, raw_data, dc_meta = latest_processed
                        stale_clients = []
                        payload = {"type": "data", "content": data, "raw": raw_data}
                        try:
                            if isinstance(raw_data, str) and len(raw_data) > 0 and raw_data[0] == 'p':
                                if isinstance(self._last_pt_mA, list) and self._last_pt_mA:
                                    payload["pt_mA"] = self._last_pt_mA[:16]
                        except Exception:
                            pass
                        try:
                            payload["devices"] = {
                                "rocket_panda": {
                                    "bb": {
                                        "l": {
                                            "enabled": self.bb_runtime["l"]["state"] != "OFF",
                                            "valve_open": bool(self.bb_runtime["l"]["press"]),
                                            "pressure": float(self.bb_runtime["l"]["pressure"]),
                                            "state": self.bb_runtime["l"]["state"],
                                            "press": bool(self.bb_runtime["l"]["press"]),
                                            "vent": bool(self.bb_runtime["l"]["vent"]),
                                        },
                                        "f": {
                                            "enabled": self.bb_runtime["f"]["state"] != "OFF",
                                            "valve_open": bool(self.bb_runtime["f"]["press"]),
                                            "pressure": float(self.bb_runtime["f"]["pressure"]),
                                            "state": self.bb_runtime["f"]["state"],
                                            "press": bool(self.bb_runtime["f"]["press"]),
                                            "vent": bool(self.bb_runtime["f"]["vent"]),
                                        },
                                    }
                                }
                            }
                        except Exception:
                            pass
                        if dc_meta is not None:
                            payload["dc"] = dc_meta
                        message = json.dumps(payload)
                        for client in list(self.clients):
                            try:
                                await client.send(message)
                                if pending_device_msgs:
                                    for evt in pending_device_msgs:
                                        await client.send(json.dumps(evt))
                            except Exception:
                                stale_clients.append(client)
                        for client in stale_clients:
                            try:
                                await client.close()
                            except Exception:
                                pass
                            if client in self.clients:
                                self.clients.remove(client)
                        self._last_ws_send_ts = now_send
                        pending_device_msgs = []
                        if self.udp_enable and self.udp_sock is not None:
                            try:
                                payload = json.dumps({"type": "data", "content": data}).encode('utf-8')
                                self.udp_sock.sendto(payload, (self.udp_group, self.udp_port))
                            except Exception:
                                pass
            except Exception:
                pass

            if processed_any and self.logging_enabled and self.log_writer is not None:
                try:
                    self._write_log_entry()
                except Exception as e:
                    logging.error(f"Failed to write log entry: {e}")

            await asyncio.sleep(0.001 if processed_any else 0.01)
    
    def start_logging(self) -> Dict:
        """Start CSV logging with timestamped filename"""
        if self.logging_enabled:
            return {"success": False, "message": "Logging already active", "filename": self.log_filename}
        
        try:
            # Create folder named with date (MMDD)
            now = datetime.now()
            date_folder = now.strftime("%m%d")
            log_dir = Path("logs") / date_folder
            log_dir.mkdir(parents=True, exist_ok=True)
            
            # Create filename with time (HHMM.csv)
            time_str = now.strftime("%H%M")
            log_path = log_dir / f"{time_str}.csv"
            
            # If file exists, append a number
            counter = 1
            while log_path.exists():
                log_path = log_dir / f"{time_str}_{counter}.csv"
                counter += 1
            
            self.log_file = open(log_path, 'w', newline='')
            self.log_writer = csv.writer(self.log_file)
            
            # Write header
            header = ['timestamp', 'elapsed_ms']
            # Add PT columns (16 channels)
            for i in range(1, 17):
                header.append(f'PT{i}_psi')
            # Add LC columns (4 channels)
            for i in range(1, 5):
                header.append(f'LC{i}_lbf')
            # Add TC columns (16 channels)
            for i in range(1, 17):
                header.append(f'TC{i}_degF')
            # Add DC current columns (16 channels, Amps)
            for i in range(1, 17):
                header.append(f'DC{i}_A')
            # Add DC boolean state columns (thresholded)
            for i in range(1, 17):
                header.append(f'DC{i}_state')
            
            self.log_writer.writerow(header)
            self.log_file.flush()
            
            self.logging_enabled = True
            self.log_filename = str(log_path)
            self.log_data_count = 0
            self.log_start_time = time.time()
            
            # Prepare command log file alongside data log
            try:
                cmd_log_path = log_dir / f"{time_str}_commands.csv"
                counter2 = 1
                while cmd_log_path.exists():
                    cmd_log_path = log_dir / f"{time_str}_commands_{counter2}.csv"
                    counter2 += 1
                self.cmd_log_file = open(cmd_log_path, 'w', newline='')
                self.cmd_log_writer = csv.writer(self.cmd_log_file)
                self.cmd_log_writer.writerow(['timestamp', 'elapsed_ms', 'command', 'success', 'result'])
                self.cmd_log_filename = str(cmd_log_path)
            except Exception as e:
                logging.error(f"Failed to init command log: {e}")
            
            logging.info(f"Started CSV logging: {self.log_filename}")
            return {"success": True, "message": "Logging started", "filename": self.log_filename}
        
        except Exception as e:
            logging.error(f"Failed to start logging: {e}")
            if self.log_file:
                self.log_file.close()
                self.log_file = None
            return {"success": False, "message": str(e)}
    
    def stop_logging(self) -> Dict:
        """Stop CSV logging"""
        if not self.logging_enabled:
            return {"success": False, "message": "Logging not active"}
        
        try:
            self.logging_enabled = False
            if self.log_file:
                self.log_file.close()
                self.log_file = None
            if self.cmd_log_file:
                try:
                    self.cmd_log_file.flush()
                    self.cmd_log_file.close()
                except Exception:
                    pass
                self.cmd_log_file = None
                self.cmd_log_writer = None
            
            logging.info(f"Stopped CSV logging. Wrote {self.log_data_count} rows to {self.log_filename}")
            filename = self.log_filename
            self.log_filename = None
            self.log_writer = None
            self.log_data_count = 0
            
            return {"success": True, "message": "Logging stopped", "rows": self.log_data_count, "filename": filename}
        
        except Exception as e:
            logging.error(f"Failed to stop logging: {e}")
            return {"success": False, "message": str(e)}
    
    def get_logging_status(self) -> Dict:
        """Get current logging status"""
        return {
            "active": self.logging_enabled,
            "filename": self.log_filename if self.logging_enabled else None,
            "rows": self.log_data_count if self.logging_enabled else 0,
            "elapsed_sec": (time.time() - self.log_start_time) if self.log_start_time else 0
        }
    
    def _write_log_entry(self):
        """Write current data to CSV log"""
        if not self.logging_enabled or self.log_writer is None:
            return
        
        with self.log_lock:
            timestamp = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
            elapsed_ms = int((time.time() - self.log_start_time) * 1000) if self.log_start_time else 0
            
            row = [timestamp, elapsed_ms]
            
            # Add PT data (16 channels)
            for i in range(1, 17):
                row.append(self.current_data['pt'].get(i, ''))
            
            # Add LC data (4 channels)
            for i in range(1, 5):
                row.append(self.current_data['lc'].get(i, ''))
            
            # Add TC data (16 channels)
            for i in range(1, 17):
                row.append(self.current_data['tc'].get(i, ''))
            
            # Add DC data (16 channels)
            for i in range(1, 17):
                row.append(self.current_data['dc'].get(i, ''))
            # Add DC thresholded state (16 channels)
            for i in range(1, 17):
                row.append(self.current_data['dc_state'].get(i, ''))
            
            self.log_writer.writerow(row)
            self.log_data_count += 1
            
            # Flush every 10 rows to ensure data is written
            if self.log_data_count % 10 == 0:
                self.log_file.flush()

    def _write_command_entry(self, command: str, success: bool | None = None, result: str = "") -> None:
        """Append a command row to the command log if available."""
        try:
            if not self.cmd_log_writer or not self.logging_enabled:
                return
            timestamp = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
            elapsed_ms = int((time.time() - self.log_start_time) * 1000) if self.log_start_time else 0
            success_cell = "" if success is None else int(bool(success))
            self.cmd_log_writer.writerow([timestamp, elapsed_ms, command, success_cell, result])
            # Flush commands more aggressively to not lose actions
            self.cmd_log_file.flush()
        except Exception as e:
            logging.error(f"Failed to write command log entry: {e}")

    def _load_presets(self):
        try:
            if self.preset_file.exists():
                with open(self.preset_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        for key in self.presets.keys():
                            value = data.get(key)
                            if isinstance(value, str) and value.strip():
                                self.presets[key] = value.strip()
        except Exception as e:
            logging.error(f"Failed to load preset sequences: {e}")

    def _save_presets(self):
        try:
            self.preset_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.preset_file, "w", encoding="utf-8") as f:
                json.dump(self.presets, f, indent=2)
        except Exception as e:
            logging.error(f"Failed to save preset sequences: {e}")

    def save_preset(self, name: str, sequence: str) -> Dict:
        preset_key = (name or "").lower()
        if preset_key not in self.presets:
            return {"success": False, "message": f"Unknown preset '{name}'"}
        if not isinstance(sequence, str) or not sequence.strip():
            return {"success": False, "message": "Sequence cannot be empty"}
        self.presets[preset_key] = sequence.strip()
        self._save_presets()
        return {"success": True, "name": preset_key, "sequence": self.presets[preset_key], "presets": self.presets}

    async def handler(self, websocket, path=None):
        self.clients.append(websocket)
        peer = None
        try:
            peer = websocket.remote_address
        except Exception:
            peer = ("?", "?")
        logging.info(f"WS client connected: {peer}, total={len(self.clients)}")
        try:
            async for message in websocket:
                try:
                    msg = json.loads(message)
                    if msg.get("action") == "connect":
                        result = await self.connect(msg["port"])
                        logging.info(f"WS connect request: port={msg.get('port')} -> {result.get('success')} {result.get('message')}")
                        await websocket.send(json.dumps(result))
                    elif msg.get("action") == "disconnect":
                        self.disconnect()
                        logging.info("WS disconnect requested by client")
                        await websocket.send(json.dumps({"success": True, "message": "Disconnected"}))
                    elif msg.get("action") == "send":
                        cmd = msg.get("command")
                        result = await self.send_command(cmd)
                        logging.info(f"WS send command: {msg.get('command')} -> {result.get('success')}")
                        await websocket.send(json.dumps(result))
                    elif msg.get("action") == "list_ports":
                        if self.test_mode:
                            # In test mode, provide fake ports
                            ports = ["TEST_PORT_1", "TEST_PORT_2"]
                        else:
                            ports = [p.device for p in serial.tools.list_ports.comports()]
                        logging.info(f"WS list_ports -> {len(ports)} ports")
                        await websocket.send(json.dumps({"ports": ports}))
                    elif msg.get("action") == "start_logging":
                        result = self.start_logging()
                        logging.info(f"WS start_logging -> {result.get('success')} {result.get('filename', '')}")
                        await websocket.send(json.dumps(result))
                    elif msg.get("action") == "stop_logging":
                        result = self.stop_logging()
                        logging.info(f"WS stop_logging -> {result.get('success')}")
                        await websocket.send(json.dumps(result))
                    elif msg.get("action") == "create_new_log":
                        # Stop current log and start a new one
                        if self.logging_enabled:
                            self.stop_logging()
                        result = self.start_logging()
                        logging.info(f"WS create_new_log -> {result.get('success')} {result.get('filename', '')}")
                        await websocket.send(json.dumps(result))
                    elif msg.get("action") == "get_logging_status":
                        result = self.get_logging_status()
                        await websocket.send(json.dumps(result))
                    elif msg.get("action") == "ping":
                        await websocket.send(json.dumps({"success": True, "message": "pong"}))
                    elif msg.get("action") == "debug_serial":
                        # Runtime toggle for raw-line printing on the server console.
                        # Body: {"action":"debug_serial", "enable":true/false, "throttle_sec":0.0}
                        try:
                            self.debug_print_serial = bool(msg.get("enable", not self.debug_print_serial))
                            if "throttle_sec" in msg:
                                self.debug_raw_throttle_sec = max(0.0, float(msg["throttle_sec"]))
                            logging.info(
                                "debug_print_serial=%s throttle=%.2fs",
                                self.debug_print_serial, self.debug_raw_throttle_sec,
                            )
                            await websocket.send(json.dumps({
                                "success": True,
                                "debug_print_serial": self.debug_print_serial,
                                "throttle_sec": self.debug_raw_throttle_sec,
                            }))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "tare_pts":
                        # Capture current raw mA snapshot and set engineering-unit tare offsets
                        try:
                            if self.pt_input_mode == "volts" and isinstance(self._last_pt_raw, list) and any(abs(v) > 0 for v in self._last_pt_raw):
                                # Compute per-channel effective shunt so current voltage -> 4mA baseline
                                new_r: List[float] = []
                                for v in (self._last_pt_raw + [0.0]*16)[:16]:
                                    try:
                                        rv = max(1.0, abs(v) / 0.004) if abs(v) > 1e-6 else self.pt_shunt_ohms
                                    except Exception:
                                        rv = self.pt_shunt_ohms
                                    new_r.append(rv)
                                self.pt_shunt_ohms_list = new_r
                                self.auto_shunt_v_ambient = False
                                # Recompute mA and offsets so displayed PSI zeros at ambient
                                mA_vals = self._to_mA(self._last_pt_raw)
                                eng_vals = self._compute_eng_from_ma_no_tare(mA_vals)
                                for i in range(min(16, len(eng_vals))):
                                    try:
                                        self.pt_offsets[i] = float(eng_vals[i])
                                    except Exception:
                                        self.pt_offsets[i] = 0.0
                                try:
                                    dbg = ', '.join(f"{r:.2f}Ω" for r in self.pt_shunt_ohms_list[:4])
                                    print(f"PT per-channel shunt (first 4): {dbg}")
                                except Exception:
                                    pass
                                await websocket.send(json.dumps({"success": True, "message": "PTs tared to 4mA baseline (per-channel shunt)"}))
                            else:
                                # Fallback: tare using current mA snapshot
                                mA_vals = self._last_pt_mA[:]
                                eng_vals = self._compute_eng_from_ma_no_tare(mA_vals)
                                for i in range(min(16, len(eng_vals))):
                                    try:
                                        self.pt_offsets[i] = float(eng_vals[i])
                                    except Exception:
                                        self.pt_offsets[i] = 0.0
                                await websocket.send(json.dumps({"success": True, "message": "PTs tared to current ambient"}))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "bb_config":
                        # Push core BB config: B<side><sp>,<db>,<wait>,<maxOpen>
                        try:
                            bus = (msg.get("bus") or "").lower()
                            if bus not in ("lox", "fuel"):
                                await websocket.send(json.dumps({"success": False, "message": "Invalid bus"}))
                                continue
                            bus_char = 'L' if bus == 'lox' else 'F'
                            sp = float(msg.get("setpoint", 200))
                            db = float(msg.get("deadband", 10))
                            wait_ms = int(msg.get("wait_ms", 500))
                            max_open_ms = int(msg.get("max_open_ms", 0))
                            cmd = f"B{bus_char}{sp:.1f},{db:.1f},{wait_ms},{max_open_ms}"
                            result = await self.send_command(cmd)
                            if result.get("success"):
                                self.bb_configs[bus]["setpoint"] = sp
                                self.bb_configs[bus]["deadband"] = db
                                self.bb_configs[bus]["wait_ms"] = wait_ms
                                self.bb_configs[bus]["max_open_ms"] = max_open_ms
                            await websocket.send(json.dumps(result))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "bb_vent_config":
                        # Push vent config: V<side><trigger>,<autoOn01>
                        try:
                            bus = (msg.get("bus") or "").lower()
                            if bus not in ("lox", "fuel"):
                                await websocket.send(json.dumps({"success": False, "message": "Invalid bus"}))
                                continue
                            bus_char = 'L' if bus == 'lox' else 'F'
                            trig = float(msg.get("trigger", 0))
                            auto_on = 1 if bool(msg.get("auto_on", False)) else 0
                            cmd = f"V{bus_char}{trig:.1f},{auto_on}"
                            result = await self.send_command(cmd)
                            if result.get("success"):
                                self.bb_configs[bus]["vent_trigger"] = trig
                                self.bb_configs[bus]["vent_auto"] = bool(auto_on)
                            await websocket.send(json.dumps(result))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "bb_mdot_config":
                        # Push mass-flow config: M<side><mdot>,<spMin>,<spMax>,<gain>,<rho>,<on01>
                        try:
                            bus = (msg.get("bus") or "").lower()
                            if bus not in ("lox", "fuel"):
                                await websocket.send(json.dumps({"success": False, "message": "Invalid bus"}))
                                continue
                            bus_char = 'L' if bus == 'lox' else 'F'
                            mdot = float(msg.get("mdot", 0))
                            sp_min = float(msg.get("sp_min", 0))
                            sp_max = float(msg.get("sp_max", 0))
                            gain = float(msg.get("gain", 0))
                            rho = float(msg.get("rho", 0))
                            enable = 1 if bool(msg.get("enable", False)) else 0
                            cmd = f"M{bus_char}{mdot:.3f},{sp_min:.3f},{sp_max:.3f},{gain:.5f},{rho:.3f},{enable}"
                            result = await self.send_command(cmd)
                            if result.get("success"):
                                self.bb_configs[bus]["mdot_target"] = mdot
                                self.bb_configs[bus]["sp_min"] = sp_min
                                self.bb_configs[bus]["sp_max"] = sp_max
                                self.bb_configs[bus]["gain"] = gain
                                self.bb_configs[bus]["rho"] = rho
                            await websocket.send(json.dumps(result))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "bb_enable":
                        # Enable/disable sustain: b<side><0|1>
                        try:
                            bus = (msg.get("bus") or "").lower()
                            if bus not in ("lox", "fuel"):
                                await websocket.send(json.dumps({"success": False, "message": "Invalid bus"}))
                                continue
                            bus_char = 'L' if bus == 'lox' else 'F'
                            enable = bool(msg.get("enable", False))
                            cmd = f"b{bus_char}{1 if enable else 0}"
                            result = await self.send_command(cmd)
                            if result.get("success"):
                                self.bb_configs[bus]["enable"] = enable
                            await websocket.send(json.dumps(result))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "bb_vent":
                        # Manual vent open/close: v<side><0|1>
                        try:
                            bus = (msg.get("bus") or "").lower()
                            if bus not in ("lox", "fuel"):
                                await websocket.send(json.dumps({"success": False, "message": "Invalid bus"}))
                                continue
                            bus_char = 'L' if bus == 'lox' else 'F'
                            open_v = bool(msg.get("open", False))
                            cmd = f"v{bus_char}{1 if open_v else 0}"
                            result = await self.send_command(cmd)
                            await websocket.send(json.dumps(result))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "bb_abort":
                        # Latched abort: x<side>
                        try:
                            bus = (msg.get("bus") or "").lower()
                            if bus not in ("lox", "fuel"):
                                await websocket.send(json.dumps({"success": False, "message": "Invalid bus"}))
                                continue
                            bus_char = 'L' if bus == 'lox' else 'F'
                            cmd = f"x{bus_char}"
                            result = await self.send_command(cmd)
                            await websocket.send(json.dumps(result))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "bb_get":
                        try:
                            await websocket.send(json.dumps({"success": True, "bb": self.bb_configs}))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "message": str(e)}))
                    elif msg.get("action") == "preset_save":
                        result = self.save_preset(msg.get("name"), msg.get("sequence"))
                        result["action"] = "preset_save"
                        await websocket.send(json.dumps(result))
                    elif msg.get("action") == "preset_load":
                        await websocket.send(json.dumps({"success": True, "action": "preset_load", "presets": self.presets}))
                    elif msg.get("action") == "save_pid_layout":
                        try:
                            layout_data = msg.get("layout", {})
                            layout_path = Path(__file__).parent / "pid_layout.json"
                            with open(layout_path, 'w') as f:
                                json.dump({"version": 1, "saved": datetime.now().isoformat(), "positions": layout_data}, f, indent=2)
                            await websocket.send(json.dumps({"success": True, "action": "save_pid_layout"}))
                            logging.info("PID layout saved")
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "action": "save_pid_layout", "message": str(e)}))
                    elif msg.get("action") == "load_pid_layout":
                        try:
                            layout_path = Path(__file__).parent / "pid_layout.json"
                            if layout_path.exists():
                                with open(layout_path, 'r') as f:
                                    data = json.load(f)
                                await websocket.send(json.dumps({"success": True, "action": "load_pid_layout", "layout": data.get("positions", {})}))
                            else:
                                await websocket.send(json.dumps({"success": True, "action": "load_pid_layout", "layout": {}}))
                        except Exception as e:
                            await websocket.send(json.dumps({"success": False, "action": "load_pid_layout", "message": str(e)}))
                except json.JSONDecodeError:
                    await websocket.send(json.dumps({"success": False, "message": "Invalid JSON"}))
        except websockets.exceptions.ConnectionClosed:
            # Normal disconnect path under browser refresh/network churn.
            pass
        finally:
            try:
                self.clients.remove(websocket)
            except ValueError:
                pass
            logging.info(f"WS client disconnected, total={len(self.clients)}")

class ReusableTCPServer(socketserver.TCPServer):
    allow_reuse_address = True

def _pick_free_port(preferred: int) -> int:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind(("0.0.0.0", preferred))
            return preferred
    except OSError:
        # Bind to 0 to get any free port
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.bind(("0.0.0.0", 0))
            return s.getsockname()[1]

def start_http_server(http_port: int):
    # Multi-root: public/ (repo root) is primary so the latest moeui.html and
    # other shared assets are served; panda/ is the fallback for legacy pages
    # and the per-board configs (pt_channels.json etc.) generated there.
    panda_root = os.path.dirname(os.path.abspath(__file__))
    public_root = os.path.abspath(os.path.join(panda_root, os.pardir, "public"))
    roots = [public_root, panda_root] if os.path.isdir(public_root) else [panda_root]
    os.chdir(roots[0])

    class _MultiRootHandler(http.server.SimpleHTTPRequestHandler):
        """Serve from public/ first, fall back to panda/ for legacy files."""
        _roots = roots

        def translate_path(self, path):
            import urllib.parse, posixpath
            path = urllib.parse.unquote(path.split('?')[0].split('#')[0])
            path = posixpath.normpath(path).lstrip('/')
            for root in self._roots:
                candidate = os.path.join(root, path) if path else root
                if os.path.isfile(candidate):
                    return candidate
            return os.path.join(self._roots[0], path)

    http_port = _pick_free_port(http_port)
    httpd = ReusableTCPServer(("0.0.0.0", http_port), _MultiRootHandler)
    threading.Thread(target=httpd.serve_forever, daemon=True).start()
    logging.info(f"HTTP server running on http://0.0.0.0:{http_port} (roots: {', '.join(roots)})")
    return httpd

async def main(test_mode=False, cli_port=None, debug_serial=False):
    # Defaults: bind on all interfaces, serve HTTP, enable UDP multicast
    ws_host = "0.0.0.0"
    ws_port = WS_PORT
    http_port = HTTP_PORT
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    # Keep UI channel configs in sync with root sensor_config.xlsx each launch.
    try:
        repo_root = Path(__file__).resolve().parent.parent
        gen_script = repo_root / "generate_configs.py"
        if gen_script.exists():
            res = subprocess.run(
                [sys.executable, str(gen_script), "--target", "panda"],
                cwd=str(repo_root),
                capture_output=True,
                text=True,
                timeout=30,
            )
            if res.returncode == 0:
                logging.info("Regenerated panda configs from sensor_config.xlsx")
            else:
                logging.warning("Config generation failed (non-fatal): %s", (res.stderr or res.stdout).strip())
        else:
            logging.warning("generate_configs.py not found at %s", gen_script)
    except Exception as e:
        logging.warning("Config auto-generation skipped: %s", e)
    processor = SerialProcessor(test_mode=test_mode, test_scenario="normal_operation", udp_enable=True, udp_group=UDP_GROUP, udp_port=UDP_PORT)
    if debug_serial:
        processor.debug_print_serial = True
        processor.debug_raw_throttle_sec = 0.0
        logging.info("Raw serial printing ENABLED (--debug-serial)")
    # Start HTTP server
    httpd = start_http_server(http_port)
    if httpd:
        logging.info("Open UI:")
        logging.info(f"  - http://localhost:{http_port}/moeui.html")
        logging.info(f"  - http://localhost:{http_port}/panda-daq-ui.html")
        logging.info(f"  - http://localhost:{http_port}/sequencer.html")
        logging.info(f"  - http://localhost:{http_port}/pid.html")
    # Start WebSocket server bound to LAN
    # In normal mode, attempt manual port if provided, otherwise auto-connect
    if not test_mode:
        try:
            connected_via_cli = False
            if cli_port:
                result = await processor.connect(cli_port)
                logging.info(f"Manual connect requested ({cli_port}): {result.get('message', '')}")
                connected_via_cli = bool(result.get('success'))
                if not connected_via_cli:
                    logging.info("Manual connect failed; attempting auto-detect...")

            if not connected_via_cli:
                all_ports = list(serial.tools.list_ports.comports())
                ports = [p.device for p in all_ports]
                if ports:
                    # Prefer likely USB serial devices
                    preferred = None
                    for p in ports:
                        if any(tag in p.lower() for tag in ["usbmodem", "usbserial", "ttyacm", "ttyusb"]):
                            preferred = p
                            break
                    chosen = preferred or ports[0]
                    await processor.connect(chosen)
                    logging.info(f"Auto-connected to serial: {chosen}")
                else:
                    logging.info("No serial ports found at startup; waiting for UI to connect later")
        except Exception as e:
            logging.error(f"Auto-connect error: {e}")

    server = await websockets.serve(
        processor.handler,
        ws_host,
        ws_port,
        ping_interval=20,
        ping_timeout=90,
        close_timeout=5,
    )
    # Ensure logging is always on once the server is up
    try:
        result = processor.start_logging()
        logging.info(f"Auto-start logging -> {result.get('success')} {result.get('filename', '')}")
    except Exception as e:
        logging.error(f"Auto-start logging failed: {e}")
    broadcast_task = asyncio.create_task(processor.broadcast_data())

    async def _cleanup():
        # Close WS server
        try:
            server.close()
            await server.wait_closed()
        except Exception:
            pass
        # Stop broadcaster
        try:
            broadcast_task.cancel()
        except Exception:
            pass
        # Shutdown HTTP
        try:
            if httpd:
                httpd.shutdown()
        except Exception:
            pass
        # Close UDP socket and serial
        try:
            if processor.udp_sock:
                processor.udp_sock.close()
        except Exception:
            pass
        try:
            processor.disconnect()
        except Exception:
            pass

    loop = asyncio.get_running_loop()

    def _signal_handler():
        # Schedule cleanup coroutine
        asyncio.ensure_future(_cleanup())
        # Stop loop after a brief delay to let cleanup run
        loop.call_later(0.1, loop.stop)

    # Register signal handlers (CTRL+C, terminate)
    try:
        loop.add_signal_handler(signal.SIGINT, _signal_handler)
        loop.add_signal_handler(signal.SIGTERM, _signal_handler)
    except NotImplementedError:
        pass

    # Ensure cleanup on interpreter exit
    atexit.register(lambda: asyncio.run(_cleanup()))
    
    mode_str = "TEST MODE" if test_mode else "NORMAL MODE"
    logging.info(f"PANDA WebSocket Bridge - {mode_str}")
    logging.info(f"WebSocket server running on ws://{ws_host}:{ws_port}")
    
    if test_mode:
        logging.info("Test telemetry enabled (scenario: normal_operation)")
    
    try:
        await asyncio.gather(server.wait_closed(), broadcast_task)
    except asyncio.CancelledError:
        pass
    finally:
        await _cleanup()

def parse_args():
    parser = argparse.ArgumentParser(description="PANDA server")
    parser.add_argument("--test", action="store_true", help="Run in test mode (simulated telemetry)")
    parser.add_argument("--port", "-p", dest="port", help="Serial port to open, e.g., COM3 or /dev/ttyUSB0")
    parser.add_argument("--debug-serial", action="store_true",
                        help="Print every raw line received from the board (very noisy)")
    return parser.parse_args()

if __name__ == "__main__":
    args = parse_args()
    asyncio.run(main(test_mode=args.test, cli_port=args.port, debug_serial=args.debug_serial))