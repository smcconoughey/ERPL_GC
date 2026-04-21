#!/usr/bin/env python3
"""
Generate JSON config files from sensor spreadsheets.

Supports two xlsx formats:
  Legacy (draco):     Sheets named PT_Sensors, LC_LoadCells, TC_Thermocouples, DC_Solenoids
                      with a 'system' column (panda / ni_daq) to split rows.
  Per-device (moe):   Sheets named after devices (rocket_panda, gse_panda, ni_daq).
                      Each sheet has a 'type' column (pt / lc / tc / dc).

Run standalone or called by start_all.sh at launch time.
"""

import json
import os
import sys

# ── Helpers ───────────────────────────────────────────────────────────

def _xlsx_path(override=None):
    if override:
        return os.path.abspath(override)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sensor_config.xlsx')


def load_xlsx(xlsx_override=None):
    import openpyxl
    path = _xlsx_path(xlsx_override)
    if not os.path.exists(path):
        print(f"[generate_configs] xlsx not found at {path}")
        return None
    print(f"[generate_configs] Loading {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _read_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(headers) if h}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for h, i in col.items():
            d[h] = row[i]
        rows.append(d)
    return rows


def _safe(val, default=''):
    return str(val) if val is not None else default

def _safef(val, default=0.0):
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default

def _safei(val, default=0):
    try:
        return int(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def _write_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
    print(f"  -> {path}")


def _detect_format(sheet_names):
    if 'PT_Sensors' in sheet_names:
        return 'legacy'
    return 'per_device'


def _split_by_type(rows):
    grouped = {'pt': [], 'lc': [], 'tc': [], 'dc': []}
    for r in rows:
        t = _safe(r.get('type')).strip().lower()
        if t in grouped:
            grouped[t].append(r)
    return grouped


# ── Channel generators (work on pre-filtered rows, no system check) ──

def _make_pt_channels(rows):
    channels = []
    for r in rows:
        ch = _safei(r.get('channel'))
        rng_min = _safef(r.get('range_min'), 0)
        rng_max = _safef(r.get('range_max'), 500)
        channels.append({
            "pt": {
                "id": ch + 1,
                "raw_index": ch + 1,
                "name": _safe(r.get('name'), f'PT{ch + 1}').upper(),
                "short_name": _safe(r.get('short_name'), f'pt{ch + 1}'),
                "units": _safe(r.get('units'), 'psi'),
                "range": [rng_min, rng_max],
            }
        })
    return {"channels": channels}


def _make_lc_channels(rows):
    channels = []
    for r in rows:
        ch = _safei(r.get('channel'))
        channels.append({
            "lc": {
                "id": ch + 1,
                "raw_index": ch + 1,
                "name": _safe(r.get('name'), f'LC{ch + 1}'),
                "short_name": _safe(r.get('short_name'), f'lc{ch + 1}'),
                "units": _safe(r.get('units'), 'lbf'),
                "range": [0, 1000],
                "calibration": {
                    "slope": _safef(r.get('cal_slope_a'), 1000000.0),
                    "slope_alt": _safef(r.get('cal_slope_b'), 1000000.0),
                    "slope_alt_label": "Cal B",
                    "offset": _safef(r.get('cal_offset'), 0.0),
                    "tare_offset": _safef(r.get('cal_tare'), 0.0),
                }
            }
        })
    return {"channels": channels}


def _make_tc_channels(rows, default_units='°C'):
    channels = []
    for r in rows:
        ch = _safei(r.get('channel'))
        tc_entry = {
            "id": ch + 1,
            "raw_index": ch + 1,
            "name": _safe(r.get('name'), f'TC{ch + 1}'),
            "short_name": _safe(r.get('short_name'), f'tc{ch + 1}'),
            "tc_type": _safe(r.get('tc_type'), 'K').upper(),
            "units": _safe(r.get('units'), default_units),
            "range": [
                _safef(r.get('cal_min_temp'), 0),
                _safef(r.get('cal_max_temp'), 2000),
            ] if r.get('cal_min_temp') is not None else [0, 2000],
            "calibration": {
                "offset": _safef(r.get('tc_cal_offset'), 0.0),
                "min_temp": _safef(r.get('cal_min_temp'), -320.0) if r.get('cal_min_temp') is not None else -320.0,
                "max_temp": _safef(r.get('cal_max_temp'), 2282.0) if r.get('cal_max_temp') is not None else 2282.0,
            }
        }
        # Chill threshold: flash UI red if value is above this (e.g. LOX line
        # must reach -280°F before liquid passes through).
        if r.get('chill_threshold') is not None:
            tc_entry["chill_threshold"] = _safef(r.get('chill_threshold'))
        channels.append({"tc": tc_entry})
    return {"channels": channels}


def _make_dc_channels(rows):
    channels = []
    for r in rows:
        ch = _safei(r.get('channel'))
        entry = {
            "id": ch + 1,
            "raw_index": ch + 1,
            "name": _safe(r.get('name'), f'DC{ch + 1}'),
            "short_name": _safe(r.get('short_name'), f'dc{ch + 1}'),
            "type": _safe(r.get('dc_type'), 'misc'),
        }
        abort_role = r.get('abort_role')
        if abort_role:
            entry["abort_role"] = str(abort_role)
        bb = r.get('bb_enabled')
        if bb and str(bb).strip().upper() == 'TRUE':
            entry["bb"] = True
        # Timer rating (minutes): UI counts up while energized, flashes red at limit.
        # Tank solenoids on MOE have a 30-min continuous-energize rating.
        if r.get('timer') is not None:
            entry["timer"] = _safei(r.get('timer'))
        channels.append({"dc": entry})
    return {"channels": channels}


def _make_bb_config(rows):
    """Generate bb_config.json from BB_Config sheet rows.

    Expected columns: bus, label, device, setpoint_psi, deadband_psi, wait_ms,
    max_open_ms, vent_trigger_psi, vent_auto, mdot_target, mdot_sp_min,
    mdot_sp_max, mdot_gain, mdot_enable, pt_channel, press_dc, vent_dc
    """
    buses = {}
    for r in rows:
        bus_key = _safe(r.get('bus'), '').strip().lower()
        if not bus_key:
            continue
        buses[bus_key] = {
            "label": _safe(r.get('label'), bus_key.upper() + ' Bus'),
            "device": _safe(r.get('device'), 'rocket_panda'),
            "setpoint_psi": _safef(r.get('setpoint_psi'), 200.0),
            "deadband_psi": _safef(r.get('deadband_psi'), 10.0),
            "wait_ms": _safei(r.get('wait_ms'), 500),
            "max_open_ms": _safei(r.get('max_open_ms'), 0),
            "vent_trigger_psi": _safef(r.get('vent_trigger_psi'), 0),
            "vent_auto": str(r.get('vent_auto', '')).strip().upper() == 'TRUE',
            "mdot_target": _safef(r.get('mdot_target'), 0),
            "mdot_sp_min": _safef(r.get('mdot_sp_min'), 0),
            "mdot_sp_max": _safef(r.get('mdot_sp_max'), 0),
            "mdot_gain": _safef(r.get('mdot_gain'), 0),
            "mdot_enable": str(r.get('mdot_enable', '')).strip().upper() == 'TRUE',
            "pt_channel": _safei(r.get('pt_channel'), 0),
            "press_dc": _safei(r.get('press_dc'), 0),
            "vent_dc": _safei(r.get('vent_dc'), 0),
            "pt_sensors": [],
            "valve": {
                "dc_id": _safei(r.get('press_dc'), 0),
                "label": _safe(r.get('valve_label'), f"DC{_safei(r.get('press_dc'), 0)}")
            }
        }
    return {"buses": buses}


def _make_nidaq_sensors(pt_rows):
    sensors = []
    for r in pt_rows:
        ch = _safei(r.get('channel'))
        sensor = {
            "channel": ch,
            "name": _safe(r.get('name'), f'PT{ch}'),
            "id": _safe(r.get('id'), f'PT{ch}E'),
            "group": _safe(r.get('group'), 'other'),
            "calibration": {
                "max_psi": _safef(r.get('cal_max_psi_a'), 10000.0),
                "zero_ma": _safef(r.get('cal_zero_ma'), 4.0),
                "units": _safe(r.get('units'), 'psi'),
            }
        }
        serial = r.get('serial')
        if serial:
            sensor["serial"] = str(serial)
        sensors.append(sensor)
    return sensors


# ── Legacy format wrappers (filter by system column) ─────────────────

def _filter_system(rows, system):
    return [r for r in rows if _safe(r.get('system')).strip().lower() == system]


def _generate_nidaq_interface(pt_rows, out_path=None):
    base = os.path.dirname(os.path.abspath(__file__))
    iface_path = out_path or os.path.join(base, 'ni_daq', 'interface_config.json')

    existing = {}
    if os.path.exists(iface_path):
        try:
            with open(iface_path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
        except Exception:
            pass

    existing["sensors"] = _make_nidaq_sensors(pt_rows)
    if "interface" not in existing:
        existing["interface"] = {
            "title": "ERPL DAQ Monitor",
            "layout": {"pid_width_percent": 50, "data_width_percent": 50},
            "theme": {"mode": "light", "high_contrast": True}
        }
    return existing, iface_path


# ── Run: legacy format ───────────────────────────────────────────────

def _run_legacy(wb, targets, out_dir, base):
    sheet_names = wb.sheetnames
    pt_rows = _read_sheet(wb, 'PT_Sensors') if 'PT_Sensors' in sheet_names else []
    lc_rows = _read_sheet(wb, 'LC_LoadCells') if 'LC_LoadCells' in sheet_names else []
    tc_rows = _read_sheet(wb, 'TC_Thermocouples') if 'TC_Thermocouples' in sheet_names else []
    dc_rows = _read_sheet(wb, 'DC_Solenoids') if 'DC_Solenoids' in sheet_names else []

    if 'panda' in targets:
        panda_dir = out_dir or os.path.join(base, 'panda', 'configs')
        panda_pt = _filter_system(pt_rows, 'panda')
        panda_lc = _filter_system(lc_rows, 'panda')
        panda_tc = _filter_system(tc_rows, 'panda')
        panda_dc = _filter_system(dc_rows, 'panda')
        _write_json(os.path.join(panda_dir, 'pt_channels.json'), _make_pt_channels(panda_pt))
        _write_json(os.path.join(panda_dir, 'lc_channels.json'), _make_lc_channels(panda_lc))
        _write_json(os.path.join(panda_dir, 'tc_channels.json'), _make_tc_channels(panda_tc))
        _write_json(os.path.join(panda_dir, 'dc_channels.json'), _make_dc_channels(panda_dc))

    if 'nidaq' in targets:
        nidaq_pt = _filter_system(pt_rows, 'ni_daq')
        nidaq_lc = _filter_system(lc_rows, 'ni_daq')
        nidaq_tc = _filter_system(tc_rows, 'ni_daq')
        iface_data, iface_path = _generate_nidaq_interface(nidaq_pt,
            os.path.join(out_dir, 'interface_config.json') if out_dir else None)
        _write_json(iface_path, iface_data)
        nidaq_dir = out_dir or os.path.join(base, 'ni_daq')
        _write_json(os.path.join(nidaq_dir, 'lc_channels.json'), _make_lc_channels(nidaq_lc))
        _write_json(os.path.join(nidaq_dir, 'tc_channels.json'), _make_tc_channels(nidaq_tc, default_units='°F'))


# ── Run: per-device format ───────────────────────────────────────────

def _run_per_device(wb, targets, out_dir, base):
    moe_dir = out_dir or os.path.join(base, 'moe', 'configs')

    for sheet_name in wb.sheetnames:
        rows = _read_sheet(wb, sheet_name)
        by_type = _split_by_type(rows)
        dev = sheet_name  # e.g. rocket_panda, gse_panda, ni_daq
        print(f"  Device sheet: {dev}  (pt={len(by_type['pt'])} lc={len(by_type['lc'])} tc={len(by_type['tc'])} dc={len(by_type['dc'])})")

        if by_type['pt']:
            _write_json(os.path.join(moe_dir, f'{dev}_pt.json'), _make_pt_channels(by_type['pt']))
        if by_type['lc']:
            _write_json(os.path.join(moe_dir, f'{dev}_lc.json'), _make_lc_channels(by_type['lc']))
        if by_type['tc']:
            units = '°F' if dev == 'ni_daq' else '°C'
            _write_json(os.path.join(moe_dir, f'{dev}_tc.json'), _make_tc_channels(by_type['tc'], default_units=units))
        if by_type['dc']:
            _write_json(os.path.join(moe_dir, f'{dev}_dc.json'), _make_dc_channels(by_type['dc']))

        # NI DAQ also gets its interface_config.json
        if dev == 'ni_daq' and by_type['pt']:
            iface_data, iface_path = _generate_nidaq_interface(by_type['pt'],
                os.path.join(out_dir, 'interface_config.json') if out_dir else None)
            _write_json(iface_path, iface_data)

    # BB_Config sheet → public/configs/bb_config.json
    if 'BB_Config' in wb.sheetnames:
        bb_rows = _read_sheet(wb, 'BB_Config')
        if bb_rows:
            bb_out = os.path.join(base, 'public', 'configs', 'bb_config.json')
            _write_json(bb_out, _make_bb_config(bb_rows))
            print(f"  BB_Config: {len(bb_rows)} bus(es)")


# ── Entry points ─────────────────────────────────────────────────────

def run(xlsx_path=None, targets=None, out_dir=None):
    """Generate JSON configs from an xlsx spreadsheet.

    Args:
        xlsx_path: Path to xlsx file. None = default sensor_config.xlsx.
        targets:   List of subsystems to generate for: 'panda', 'nidaq', 'moe'.
                   None = all (panda + nidaq, legacy behavior).
        out_dir:   Override output directory for generated configs.
                   None = default locations per subsystem.
    """
    wb = load_xlsx(xlsx_path)
    if wb is None:
        return False

    base = os.path.dirname(os.path.abspath(__file__))
    fmt = _detect_format(wb.sheetnames)

    if targets is None:
        targets = ['panda', 'nidaq'] if fmt == 'legacy' else ['moe']

    print(f"[generate_configs] Format: {fmt}, targets: {', '.join(targets)}")

    if fmt == 'legacy':
        _run_legacy(wb, targets, out_dir, base)
    else:
        _run_per_device(wb, targets, out_dir, base)

    wb.close()
    print("[generate_configs] Done.")
    return True


def parse_args():
    import argparse
    parser = argparse.ArgumentParser(
        description="Generate JSON configs from sensor spreadsheet")
    parser.add_argument('--xlsx', default=None,
        help='Path to xlsx config sheet (default: sensor_config.xlsx)')
    parser.add_argument('--target', action='append', dest='targets',
        choices=['panda', 'nidaq', 'moe'],
        help='Subsystem to generate for (repeatable). Default: auto-detect.')
    parser.add_argument('--out-dir', default=None,
        help='Override output directory for all generated configs')
    return parser.parse_args()


if __name__ == '__main__':
    args = parse_args()
    success = run(xlsx_path=args.xlsx, targets=args.targets, out_dir=args.out_dir)
    sys.exit(0 if success else 1)
