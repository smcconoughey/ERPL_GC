# NI-9211 4-channel thermocouple configuration
import nidaqmx
from nidaqmx.constants import AcquisitionType, ThermocoupleType, TemperatureUnits, CJCSource
from typing import Dict, List, Any
from .base_device import BaseDevice
import nidaqmx.system
import logging
import json
from pathlib import Path

logger = logging.getLogger(__name__)

class TCCard(BaseDevice):
    """NI-9211 4-channel thermocouple input configuration.
    
    The NI-9211 is a slow, high-precision TC module (~14 S/s aggregate).
    It does NOT support hardware sample clock timing - must use on-demand reads.
    """
    
    def __init__(self, chassis: str, module_slot: int):
        super().__init__(chassis, module_slot)
        self.device_name = self.module_name
        self.product_type = "Unknown"
        self.channel_count = 4
        # NI-9211 max is ~14 S/s total across all channels (~3.5 Hz per channel)
        # We use on-demand reads, so this is just for timing calculations
        self.sample_rate = 1.0
        self.samples_per_channel = 1
        
        # Flag to indicate this device uses on-demand reads (no hardware timing)
        self.on_demand = True
        
        # Autodetect the NI-9211 module on the chassis
        try:
            system = nidaqmx.system.System.local()
            for dev in system.devices:
                if dev.name.startswith(chassis) and "9211" in dev.product_type:
                    self.module_name = dev.name
                    try:
                        self.module_slot = int(dev.name.split('Mod')[-1])
                    except Exception:
                        pass
                    self.device_name = self.module_name
                    self.product_type = dev.product_type
                    logger.info(f"NI-9211 autodetected: {self.module_name}")
                    break
        except Exception as e:
            logger.warning(f"NI-9211 autodetect failed: {e}")
        
        self.tc_config = {}
        self.tare_offsets = {}
        self._load_tc_config()
    
    def _load_tc_config(self) -> None:
        """Load TC channel configuration from sensor_config.xlsx"""
        try:
            import openpyxl
            
            base = Path(__file__).parent.parent
            xlsx_path = base / '..' / 'sensor_config.xlsx'
            if not xlsx_path.exists():
                xlsx_path = base / 'sensor_config.xlsx'
            
            wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
            ws = wb['TC_Thermocouples']
            
            headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            col = {h: i for i, h in enumerate(headers) if h}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                system = row[col.get('system', 0)]
                if not system or str(system).strip().lower() != 'ni_daq':
                    continue
                
                try:
                    channel = int(row[col['channel']])
                except (ValueError, KeyError, TypeError):
                    continue
                
                if channel >= self.channel_count:
                    continue
                
                min_temp = row[col.get('cal_min_temp', -1)] if col.get('cal_min_temp') is not None else None
                max_temp = row[col.get('cal_max_temp', -1)] if col.get('cal_max_temp') is not None else None
                tc_offset = row[col.get('tc_cal_offset', -1)] if col.get('tc_cal_offset') is not None else None
                tc_type = row[col.get('tc_type', -1)] if col.get('tc_type') is not None else None
                
                self.tc_config[channel] = {
                    'id': channel + 1,
                    'name': str(row[col.get('name', -1)] or f'TC{channel}'),
                    'short_name': str(row[col.get('short_name', -1)] or f'TC{channel}'),
                    'tc_type': str(tc_type or 'K').upper(),
                    'units': str(row[col.get('units', -1)] or '°F'),
                    'calibration': {
                        'offset': float(tc_offset) if tc_offset is not None else 0.0,
                        'min_temp': float(min_temp) if min_temp is not None else -320.0,
                        'max_temp': float(max_temp) if max_temp is not None else 2282.0,
                    }
                }
            
            wb.close()
            logger.info(f"Loaded TC config for {len(self.tc_config)} channels from sensor_config.xlsx")
        except Exception as e:
            logger.warning(f"Could not load TC config from xlsx: {e}")
    
    def _get_tc_type(self, channel: int) -> ThermocoupleType:
        """Get thermocouple type for a channel from config."""
        if channel in self.tc_config:
            tc_type_str = self.tc_config[channel].get('tc_type', 'K').upper()
            tc_types = {
                'K': ThermocoupleType.K,
                'J': ThermocoupleType.J,
                'T': ThermocoupleType.T,
                'E': ThermocoupleType.E,
                'R': ThermocoupleType.R,
                'S': ThermocoupleType.S,
                'B': ThermocoupleType.B,
                'N': ThermocoupleType.N,
            }
            return tc_types.get(tc_type_str, ThermocoupleType.K)
        return ThermocoupleType.K
    
    def _get_sensor_info(self, channel: int) -> Dict[str, str]:
        """Get sensor info for a channel."""
        if channel in self.tc_config:
            tc = self.tc_config[channel]
            return {
                'name': tc.get('name', f'TC{channel}'),
                'id': tc.get('short_name', f'TC{channel}'),
                'group': 'thermocouple'
            }
        return {
            'name': f'TC{channel}',
            'id': f'TC{channel}',
            'group': 'thermocouple'
        }
    
    def configure_channels(self, task: nidaqmx.Task) -> None:
        """Configure NI-9211 thermocouple channels.
        
        The NI-9211 uses:
        - Built-in CJC (Cold Junction Compensation) sensor
        - Autozero for drift correction
        - 24-bit ADC for high precision
        """
        for ch_idx in range(self.channel_count):
            channel_name = f"{self.device_name}/ai{ch_idx}"
            tc_type = self._get_tc_type(ch_idx)
            
            # Get min/max from config or use defaults for cryogenic + hot range
            # Type K polynomial scaling limit is -328°F, so default to -320°F
            if ch_idx in self.tc_config:
                cal = self.tc_config[ch_idx].get('calibration', {})
                min_val = cal.get('min_temp', -320.0)
                max_val = cal.get('max_temp', 2282.0)
            else:
                min_val = -320.0
                max_val = 2282.0
            
            task.ai_channels.add_ai_thrmcpl_chan(
                physical_channel=channel_name,
                min_val=min_val,
                max_val=max_val,
                units=TemperatureUnits.DEG_F,
                thermocouple_type=tc_type,
                cjc_source=CJCSource.BUILT_IN,
                cjc_val=77.0,  # Fallback CJC value (room temp °F)
            )
        
        logger.info(f"NI-9211: Configured {self.channel_count} thermocouple channels")
    
    def configure_timing(self, task: nidaqmx.Task) -> None:
        """Configure timing for NI-9211.
        
        IMPORTANT: The NI-9211 does NOT support hardware sample clock timing.
        It must use on-demand (software-timed) reads only.
        We don't configure any timing here - reads will be on-demand.
        """
        # NI-9211 does NOT support cfg_samp_clk_timing
        # Leave timing unconfigured for on-demand reads
        logger.info("NI-9211: Using on-demand (software-timed) reads - no hardware timing configured")
    
    @property
    def device_info(self) -> Dict[str, Any]:
        return {
            'device_type': 'TC Card (NI-9211)',
            'module': self.module_name,
            'channels': self.channel_count,
            'sample_rate': self.sample_rate,
            'on_demand': True
        }
    
    def convert_temperature(self, raw_temp: float, channel: int) -> float:
        """Apply any calibration offset to temperature reading."""
        if channel in self.tc_config:
            cal = self.tc_config[channel].get('calibration', {})
            offset = cal.get('offset', 0.0)
            raw_temp += offset
        
        tare = self.tare_offsets.get(channel, 0.0)
        return raw_temp - tare
    
    def tare(self, baseline_raw_data: List[List[float]]) -> None:
        """Tare TC channels to current reading (useful for differential measurements)."""
        for ch_idx in range(min(self.channel_count, len(baseline_raw_data))):
            samples = baseline_raw_data[ch_idx] if isinstance(baseline_raw_data[ch_idx], list) else []
            if not samples:
                continue
            avg_temp = sum(samples) / len(samples)
            self.tare_offsets[ch_idx] = avg_temp
            logger.info(f"Tared TC{ch_idx}: offset = {avg_temp:.2f}°F")
    
    def tare_channels(self, channels: List[int], baseline_raw_data: List[List[float]]) -> None:
        """Tare only specified TC channels."""
        if not channels:
            return
        for ch_idx in channels:
            if ch_idx < 0 or ch_idx >= min(self.channel_count, len(baseline_raw_data)):
                continue
            samples = baseline_raw_data[ch_idx] if isinstance(baseline_raw_data[ch_idx], list) else []
            if not samples:
                continue
            avg_temp = sum(samples) / len(samples)
            self.tare_offsets[ch_idx] = avg_temp
            logger.info(f"Tared TC{ch_idx} via targeted request: offset = {avg_temp:.2f}°F")
    
    def process_data(self, raw_data: List[List[float]]) -> Dict[str, Any]:
        """Process raw temperature data into structured format."""
        result: List[Dict[str, Any]] = []
        per_sample_records: List[List[Dict[str, Any]]] = []
        
        if not isinstance(raw_data, list):
            return {'channels': result, 'samples': []}
        
        for ch_idx in range(min(self.channel_count, len(raw_data))):
            samples = raw_data[ch_idx] if isinstance(raw_data[ch_idx], list) else []
            if not samples:
                avg_temp = 0.0
            else:
                avg_temp = sum(samples) / len(samples)
            
            info = self._get_sensor_info(ch_idx)
            
            # Check for open/disconnected TC
            # NaN or extreme out-of-range indicates open TC
            # Note: -350°F is valid for LOX, so threshold must be lower
            status = 'ok'
            import math
            if math.isnan(avg_temp) or avg_temp < -450 or avg_temp > 3500:
                status = 'disconnected'
            
            temp_f = self.convert_temperature(avg_temp, ch_idx) if status == 'ok' else 0.0
            
            result.append({
                'channel': ch_idx,
                'name': info['name'],
                'id': info['id'],
                'group': info['group'],
                'temp_raw': round(avg_temp, 2) if status == 'ok' else 0.0,
                'temp_f': round(temp_f, 2),
                'status': status,
                'units': '°F'
            })
            
            # Build per-sample records for CSV logging
            for sample_idx, sample_val in enumerate(samples):
                while len(per_sample_records) <= sample_idx:
                    per_sample_records.append([])
                
                sample_status = 'ok'
                if math.isnan(sample_val) or sample_val < -450 or sample_val > 3500:
                    sample_status = 'disconnected'
                
                sample_temp = self.convert_temperature(sample_val, ch_idx) if sample_status == 'ok' else 0.0
                
                per_sample_records[sample_idx].append({
                    'channel': ch_idx,
                    'temp_raw': round(sample_val, 2) if sample_status == 'ok' else 0.0,
                    'temp_f': round(sample_temp, 2),
                    'status': sample_status,
                })
        
        return {'channels': result, 'samples': per_sample_records}
