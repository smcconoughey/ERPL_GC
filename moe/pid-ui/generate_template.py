#!/usr/bin/env python3
"""Generate the MOE P&ID template XLSX file.

Creates an Excel workbook with:
  - PID sheet: grid layout of the MOE propulsion system
  - Config sheet: element-to-device channel mapping
  - Legend sheet: quick reference for cell symbols

Usage:
    pip install openpyxl
    python generate_template.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Color fills for fluid types ─────────────────────────────────────
LOX    = PatternFill(start_color="3182CE", end_color="3182CE", fill_type="solid")
FUEL   = PatternFill(start_color="ED8936", end_color="ED8936", fill_type="solid")
GN2    = PatternFill(start_color="38A169", end_color="38A169", fill_type="solid")
PURGE  = PatternFill(start_color="9F7AEA", end_color="9F7AEA", fill_type="solid")
HEADER = PatternFill(start_color="1A202C", end_color="1A202C", fill_type="solid")
LIGHT  = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=9)
CELL_FONT  = Font(bold=True, size=9, name="Consolas")
LABEL_FONT = Font(bold=True, size=8, name="Consolas", color="718096")
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def set_cell(ws, row, col, value, fill=None, font=None):
    """Set a cell value with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or CELL_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


def create_pid_sheet(wb):
    """Create the main P&ID layout sheet."""
    ws = wb.active
    ws.title = "PID"

    # Set cell dimensions for roughly square cells
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, 30):
        ws.row_dimensions[row].height = 42

    # Apply thin borders and center alignment to all cells in the grid
    for r in range(1, 26):
        for c in range(1, 18):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.font = CELL_FONT

    # ── Layout the MOE P&ID ──────────────────────────────────────────
    # Row 1: Labels
    set_cell(ws, 1, 1, "MOE P&ID", font=Font(bold=True, size=12, color="E53E3E"))

    # Row 2: GN2 supply at top center
    set_cell(ws, 2, 9, "GN2", fill=GN2, font=WHITE_FONT)

    # Row 3: Pipe down from GN2
    set_cell(ws, 3, 9, "|", fill=GN2, font=WHITE_FONT)

    # Row 4: Pressurant PT
    set_cell(ws, 4, 9, "|,PT10", fill=GN2, font=WHITE_FONT)

    # Row 5: Junction splitting L/R to tanks
    set_cell(ws, 5, 5, "-", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 6, "-", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 7, "-", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 8, "-", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 9, "+", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 10, "-", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 11, "-", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 12, "-", fill=GN2, font=WHITE_FONT)
    set_cell(ws, 5, 13, "-", fill=GN2, font=WHITE_FONT)

    # Row 6: Pressurization valves
    set_cell(ws, 6, 5, "PB6", fill=GN2, font=WHITE_FONT)  # LOX pressurize
    set_cell(ws, 6, 13, "PB7", fill=GN2, font=WHITE_FONT)  # Fuel pressurize

    # Row 7: Pipe down to tanks
    set_cell(ws, 7, 5, "|", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 7, 13, "|", fill=FUEL, font=WHITE_FONT)

    # Row 8: Vent valves (left of LOX, right of Fuel)
    set_cell(ws, 8, 3, "PB4")   # LOX vent valve
    set_cell(ws, 8, 4, "-", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 8, 5, "+", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 8, 13, "+", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 8, 14, "-", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 8, 15, "PB5")  # Fuel vent valve

    # Row 9: PT on vent lines + pipe to tank
    set_cell(ws, 9, 3, "|,PV1", font=LABEL_FONT)  # LOX vent PT
    set_cell(ws, 9, 5, "LOX_TANK", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 9, 13, "FUEL_TANK", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 9, 15, "|,PV2", font=LABEL_FONT)  # Fuel vent PT

    # Row 10: Tank PT sensors
    set_cell(ws, 10, 4, "-", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 10, 5, "+", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 10, 6, "PT1", fill=LOX, font=WHITE_FONT)  # LOX tank PT
    set_cell(ws, 10, 12, "PT2", fill=FUEL, font=WHITE_FONT)  # Fuel tank PT
    set_cell(ws, 10, 13, "+", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 10, 14, "-", fill=FUEL, font=WHITE_FONT)

    # Row 11: Fill valves (GSE side)
    set_cell(ws, 11, 3, "PB11")  # LOX fill (GSE)
    set_cell(ws, 11, 4, "-", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 11, 5, "+", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 11, 13, "+", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 11, 14, "-", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 11, 15, "PB12")  # Fuel fill (GSE)

    # Row 12: Pipe down from tanks
    set_cell(ws, 12, 5, "|", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 12, 13, "|", fill=FUEL, font=WHITE_FONT)

    # Row 13: Main run valves
    set_cell(ws, 13, 5, "PB1", fill=LOX, font=WHITE_FONT)   # LOX main
    set_cell(ws, 13, 13, "PB2", fill=FUEL, font=WHITE_FONT)  # Fuel main

    # Row 14: Pipe down
    set_cell(ws, 14, 5, "|", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 14, 13, "|", fill=FUEL, font=WHITE_FONT)

    # Row 15: Injector PTs
    set_cell(ws, 15, 5, "|,PT3", fill=LOX, font=WHITE_FONT)   # LOX injector
    set_cell(ws, 15, 13, "|,PT4", fill=FUEL, font=WHITE_FONT)  # Fuel injector

    # Row 16: Pipe down
    set_cell(ws, 16, 5, "|", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 16, 13, "|", fill=FUEL, font=WHITE_FONT)

    # Row 17: Inlet PTs
    set_cell(ws, 17, 5, "|,PT6", fill=LOX, font=WHITE_FONT)   # LOX inlet
    set_cell(ws, 17, 13, "|,PT7", fill=FUEL, font=WHITE_FONT)  # Fuel inlet

    # Row 18: Pipes converging to engine
    set_cell(ws, 18, 5, "|", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 18, 6, "-", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 18, 7, "-", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 18, 8, "-", fill=LOX, font=WHITE_FONT)
    set_cell(ws, 18, 9, "+")
    set_cell(ws, 18, 10, "-", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 18, 11, "-", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 18, 12, "-", fill=FUEL, font=WHITE_FONT)
    set_cell(ws, 18, 13, "|", fill=FUEL, font=WHITE_FONT)

    # Row 19: Igniter + double pipe into engine
    set_cell(ws, 19, 7, "PB3")  # Igniter
    set_cell(ws, 19, 9, "||")

    # Row 20: Engine
    set_cell(ws, 20, 9, "Engine")

    # Row 21: Chamber PT
    set_cell(ws, 21, 9, "|,PT5")

    # Row 22: Thrust LC
    set_cell(ws, 22, 9, "LC1")

    return ws


def create_config_sheet(wb):
    """Create the Config sheet mapping elements to data channels."""
    ws = wb.create_sheet("Config")

    headers = ["Element", "Device", "Type", "Channel", "Units"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = WHITE_FONT
        cell.fill = HEADER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Rocket PT channels
    mappings = [
        ("PT1",  "rocket_panda", "pt", "1",  "psi"),   # LOX tank
        ("PT2",  "rocket_panda", "pt", "2",  "psi"),   # Fuel tank
        ("PT3",  "rocket_panda", "pt", "3",  "psi"),   # LOX injector
        ("PT4",  "rocket_panda", "pt", "4",  "psi"),   # Fuel injector
        ("PT5",  "rocket_panda", "pt", "5",  "psi"),   # Chamber
        ("PT6",  "rocket_panda", "pt", "6",  "psi"),   # LOX inlet
        ("PT7",  "rocket_panda", "pt", "7",  "psi"),   # Fuel inlet
        ("PV1",  "rocket_panda", "pt", "8",  "psi"),   # LOX vent
        ("PV2",  "rocket_panda", "pt", "9",  "psi"),   # Fuel vent
        ("PT10", "rocket_panda", "pt", "10", "psi"),   # Pressurant
        # Rocket DC channels (valves)
        ("PB1",  "rocket_panda", "dc", "1",  ""),      # LOX main
        ("PB2",  "rocket_panda", "dc", "2",  ""),      # Fuel main
        ("PB3",  "rocket_panda", "dc", "3",  ""),      # Igniter
        ("PB4",  "rocket_panda", "dc", "4",  ""),      # LOX vent
        ("PB5",  "rocket_panda", "dc", "5",  ""),      # Fuel vent
        ("PB6",  "rocket_panda", "dc", "6",  ""),      # LOX pressurize
        ("PB7",  "rocket_panda", "dc", "7",  ""),      # Fuel pressurize
        # GSE DC channels
        ("PB11", "gse_panda",    "dc", "1",  ""),      # LOX fill
        ("PB12", "gse_panda",    "dc", "2",  ""),      # Fuel fill
        # Load cells
        ("LC1",  "rocket_panda", "lc", "1",  "lbf"),   # Thrust
    ]

    for i, (elem, device, typ, ch, units) in enumerate(mappings, 2):
        ws.cell(row=i, column=1, value=elem).font = Font(bold=True, size=10, name="Consolas")
        ws.cell(row=i, column=2, value=device).font = Font(size=10, name="Consolas")
        ws.cell(row=i, column=3, value=typ).font = Font(size=10, name="Consolas")
        ws.cell(row=i, column=4, value=ch).font = Font(size=10, name="Consolas")
        ws.cell(row=i, column=5, value=units).font = Font(size=10, name="Consolas")
        for c in range(1, 6):
            ws.cell(row=i, column=c).alignment = CENTER
            ws.cell(row=i, column=c).border = THIN_BORDER
        if i % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=i, column=c).fill = LIGHT

    return ws


def create_legend_sheet(wb):
    """Create the Legend reference sheet."""
    ws = wb.create_sheet("Legend")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45

    headers = ["Symbol", "Description"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = WHITE_FONT
        cell.fill = HEADER
        cell.alignment = CENTER

    entries = [
        ("|",         "Vertical pipe"),
        ("-",         "Horizontal pipe"),
        ("+",         "Junction / corner (auto-connects to neighbors)"),
        ("||",        "Double vertical pipe (two parallel lines)"),
        ("--",        "Double horizontal pipe"),
        ("PT#",       "Pressure transducer (circle, shows live psi)"),
        ("TC#",       "Thermocouple (shows live temperature)"),
        ("LC#",       "Load cell (shows live force/weight)"),
        ("PV#",       "Vent pressure sensor"),
        ("PI#",       "Injector pressure sensor"),
        ("PB#",       "Ball valve (bowtie shape, clickable when armed)"),
        ("SV#",       "Solenoid valve"),
        ("|,PT5",     "Vertical pipe with PT5 sensor branch"),
        ("-,PB3",     "Horizontal pipe with inline valve"),
        ("Engine",    "Engine / combustion chamber (triangle)"),
        ("LOX_TANK",  "LOX propellant tank (blue rectangle)"),
        ("FUEL_TANK", "Fuel propellant tank (orange rectangle)"),
        ("GN2",       "GN2 pressurant supply (green diamond)"),
        ("",          ""),
        ("COLORS:",   "Set cell FILL COLOR in Excel for fluid type:"),
        ("Blue",      "LOX lines (#3182CE or similar blue)"),
        ("Orange",    "Fuel lines (#ED8936 or similar orange)"),
        ("Green",     "GN2 / pressurant lines (#38A169 or similar green)"),
        ("Purple",    "Purge lines (#9F7AEA or similar purple)"),
        ("No fill",   "Default / neutral (gray pipe)"),
    ]

    for i, (sym, desc) in enumerate(entries, 2):
        ws.cell(row=i, column=1, value=sym).font = Font(bold=True, size=10, name="Consolas")
        ws.cell(row=i, column=1, value=sym).alignment = CENTER
        ws.cell(row=i, column=2, value=desc).font = Font(size=10)
        ws.cell(row=i, column=2, value=desc).alignment = Alignment(vertical="center")
        for c in range(1, 3):
            ws.cell(row=i, column=c).border = THIN_BORDER

    # Add color swatches
    color_rows = {22: LOX, 23: FUEL, 24: GN2, 25: PURGE}
    for row, fill in color_rows.items():
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).font = Font(bold=True, size=10, name="Consolas", color="FFFFFF")

    return ws


def main():
    wb = openpyxl.Workbook()

    create_pid_sheet(wb)
    # No Config sheet needed - element mapping comes from moe/configs/ at runtime
    create_legend_sheet(wb)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, "public", "moe_pid_layout.xlsx")
    wb.save(out_path)
    print(f"Template saved to: {out_path}")

    # Also save a copy in the pid-ui root for easy access
    root_path = os.path.join(out_dir, "moe_pid_layout.xlsx")
    wb.save(root_path)
    print(f"Copy saved to: {root_path}")


if __name__ == "__main__":
    main()
